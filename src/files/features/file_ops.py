"""
Core file and folder CRUD operations for the Files Agent.

All functions return {"status": "success"|"error", ...} so the orchestrator
can report errors consistently.
"""
from __future__ import annotations

import logging
import os
import re
import shutil
import string
import subprocess
import platform
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any, Dict, List, Optional

from ..files_service import resolve_path, _file_dict, _is_safe_path, send_to_recycle_bin

logger = logging.getLogger("files_agent")


def _human_size(b: int) -> str:
    """Convert bytes to human-readable string (B / KB / MB / GB / TB)."""
    for unit in ("B", "KB", "MB", "GB", "TB"):
        if b < 1024:
            return f"{b:.1f} {unit}"
        b /= 1024
    return f"{b:.1f} PB"


def list_directory(path: str, show_hidden: bool = False, limit: int = 200) -> Dict[str, Any]:
    """
    List files and subfolders at *path*.

    Args:
        path:        Directory to list. Accepts ~ and env vars.
        show_hidden: Include files/folders starting with '.' or hidden attribute (Windows).
        limit:       Maximum number of entries to return.
    """
    try:
        p = resolve_path(path)
        if not p.exists():
            return {"status": "error", "message": f"Path does not exist: {p}"}
        if not p.is_dir():
            return {"status": "error", "message": f"Not a directory: {p}"}

        entries = []
        for child in sorted(p.iterdir(), key=lambda x: (x.is_file(), x.name.lower())):
            if not show_hidden and child.name.startswith("."):
                continue
            entries.append(_file_dict(child))
            if len(entries) >= limit:
                break

        folders = [e for e in entries if e["type"] == "folder"]
        files   = [e for e in entries if e["type"] == "file"]
        return {
            "status": "success",
            "path": str(p),
            "total_entries": len(entries),
            "folders": len(folders),
            "files": len(files),
            "entries": entries,
        }
    except PermissionError:
        return {"status": "error", "message": f"Permission denied: {path}"}
    except Exception as exc:
        logger.error("list_directory failed: %s", exc)
        return {"status": "error", "message": str(exc)}


def get_file_info(path: str) -> Dict[str, Any]:
    """
    Get detailed metadata for a file or folder.

    Args:
        path: Path to the file or folder.
    """
    try:
        p = resolve_path(path)
        if not p.exists():
            return {"status": "error", "message": f"Path does not exist: {p}"}
        info = _file_dict(p, include_hash=p.is_file())
        info["status"] = "success"
        info["absolute_path"] = str(p.resolve())
        info["parent"] = str(p.parent)
        return info
    except Exception as exc:
        logger.error("get_file_info failed: %s", exc)
        return {"status": "error", "message": str(exc)}


def copy_file(source: str, destination: str) -> Dict[str, Any]:
    """
    Copy a file or folder tree from *source* to *destination*.

    Args:
        source:      Path to copy from.
        destination: Path to copy to. If destination is a directory, the source
                     is copied inside it.
    """
    try:
        src = resolve_path(source)
        dst = resolve_path(destination)
        if not src.exists():
            return {"status": "error", "message": f"Source does not exist: {src}"}
        if src.is_dir():
            out = shutil.copytree(src, dst / src.name if dst.is_dir() else dst)
        else:
            out = shutil.copy2(src, dst)
        return {"status": "success", "message": f"Copied to {out}", "destination": str(out)}
    except Exception as exc:
        logger.error("copy_file failed: %s", exc)
        return {"status": "error", "message": str(exc)}


def collect_files_to_folder(file_paths: List[str], destination: str) -> Dict[str, Any]:
    """
    Copy a list of files (from potentially different locations) into a single
    destination folder, creating it if needed.

    Use this when you have a list of file paths from a search result and want
    to gather them in one place before zipping or sharing.

    Args:
        file_paths:  List of absolute paths to files (or folders) to copy.
        destination: Absolute path of the folder to copy everything into.
    """
    try:
        dst = resolve_path(destination)
        dst.mkdir(parents=True, exist_ok=True)

        copied: List[str] = []
        skipped: List[str] = []
        for src_str in file_paths:
            src = resolve_path(src_str)
            if not src.exists():
                skipped.append(str(src))
                continue
            target = dst / src.name
            # Avoid name collision — append a counter suffix
            if target.exists():
                stem, suffix = src.stem, src.suffix
                for i in range(1, 1000):
                    target = dst / f"{stem}_{i}{suffix}"
                    if not target.exists():
                        break
            if src.is_dir():
                shutil.copytree(str(src), str(target), dirs_exist_ok=True)
            else:
                shutil.copy2(str(src), str(target))
            copied.append(str(target))

        _log_operation("copy", str(dst), len(copied))
        return {
            "status": "success",
            "destination": str(dst),
            "file_path": str(dst),   # convenience alias for zip_folder
            "copied_count": len(copied),
            "copied": copied,
            "skipped": skipped,
            "message": (
                f"Collected {len(copied)} item(s) into '{dst}'."
                + (f" Skipped {len(skipped)} missing path(s)." if skipped else "")
            ),
        }
    except Exception as exc:
        logger.error("collect_files_to_folder failed: %s", exc)
        return {"status": "error", "message": str(exc)}


def move_file(source: str, destination: str) -> Dict[str, Any]:
    """
    Move or rename a file or folder.

    Args:
        source:      Path to move from.
        destination: New path (or destination directory).
    """
    try:
        src = resolve_path(source)
        dst = resolve_path(destination)
        if not src.exists():
            return {"status": "error", "message": f"Source does not exist: {src}"}
        if not _is_safe_path(src):
            return {"status": "error", "message": "Cannot move from a protected system path."}
        out = shutil.move(str(src), str(dst))
        return {"status": "success", "message": f"Moved to {out}", "destination": str(out)}
    except Exception as exc:
        logger.error("move_file failed: %s", exc)
        return {"status": "error", "message": str(exc)}


def delete_file(path: str, permanent: bool = False) -> Dict[str, Any]:
    """
    Delete a file or folder.

    Args:
        path:      Path to delete.
        permanent: If True, bypass Recycle Bin. Default False (safer).
    """
    try:
        p = resolve_path(path)
        if not p.exists():
            return {"status": "error", "message": f"Path does not exist: {p}"}
        if not _is_safe_path(p):
            return {"status": "error", "message": "Cannot delete a protected system path."}
        if permanent:
            if p.is_dir():
                shutil.rmtree(p)
            else:
                p.unlink()
            msg = f"Permanently deleted: {p}"
        else:
            send_to_recycle_bin(p)
            msg = f"Moved to Recycle Bin: {p}"
        return {"status": "success", "message": msg}
    except Exception as exc:
        logger.error("delete_file failed: %s", exc)
        return {"status": "error", "message": str(exc)}


def create_folder(path: str) -> Dict[str, Any]:
    """
    Create a directory and any missing parent directories.

    Args:
        path: Full path of the new folder to create.
    """
    try:
        p = resolve_path(path)
        if p.exists():
            return {"status": "success", "message": f"Folder already exists: {p}", "path": str(p)}
        p.mkdir(parents=True, exist_ok=True)
        return {"status": "success", "message": f"Created folder: {p}", "path": str(p)}
    except Exception as exc:
        logger.error("create_folder failed: %s", exc)
        return {"status": "error", "message": str(exc)}


def rename_file(path: str, new_name: str) -> Dict[str, Any]:
    """
    Rename a file or folder (keeping it in the same directory).

    Args:
        path:     Current path of the file or folder.
        new_name: New name only (not a full path).
    """
    try:
        p = resolve_path(path)
        if not p.exists():
            return {"status": "error", "message": f"Path does not exist: {p}"}
        new_path = p.parent / new_name
        if new_path.exists():
            return {"status": "error", "message": f"Destination already exists: {new_path}"}
        p.rename(new_path)
        return {"status": "success", "message": f"Renamed to {new_path.name}", "new_path": str(new_path)}
    except Exception as exc:
        logger.error("rename_file failed: %s", exc)
        return {"status": "error", "message": str(exc)}


def open_file(path: str) -> Dict[str, Any]:
    """
    Open a file with its default application (Windows: os.startfile, others: xdg-open/open).

    Args:
        path: Path to the file to open.
    """
    try:
        p = resolve_path(path)
        if not p.exists():
            return {"status": "error", "message": f"File does not exist: {p}"}
        if not p.is_file():
            return {"status": "error", "message": f"Not a file: {p}"}
        system = platform.system()
        if system == "Windows":
            os.startfile(str(p))
        elif system == "Darwin":
            subprocess.Popen(["open", str(p)])
        else:
            subprocess.Popen(["xdg-open", str(p)])
        return {"status": "success", "message": f"Opened {p.name} with default application"}
    except Exception as exc:
        logger.error("open_file failed: %s", exc)
        return {"status": "error", "message": str(exc)}


_SYSTEM_FOLDERS = frozenset({
    "Windows", "Program Files", "Program Files (x86)", "ProgramData",
    "$Recycle.Bin", "System Volume Information", "Recovery", "boot",
    "bootmgr", "hiberfil.sys", "pagefile.sys", "swapfile.sys",
    "dell", "dell.sdr", "e-logo", "inetpub", "Intel", "OneDriveTemp",
    "tmp", "Documents and Settings", "MSOCache", "PerfLogs",
    "AppData", "Application Data", "Local Settings", "Cookies",
    "NetHood", "PrintHood", "Recent", "SendTo", "Start Menu", "Templates",
    "anaconda3",
})


def _scan_dir_recursive(
    path: Path,
    depth: int,
    include_hidden: bool,
) -> List[Dict[str, Any]]:
    """Recursively scan *path* up to *depth* levels. depth=1 → immediate children only."""
    entries: List[Dict[str, Any]] = []
    try:
        for child in sorted(path.iterdir(), key=lambda x: (x.is_file(), x.name.lower())):
            if not include_hidden and (child.name.startswith(".") or child.name.startswith("$")):
                continue
            entry: Dict[str, Any] = {
                "name": child.name,
                "type": "folder" if child.is_dir() else "file",
            }
            if child.is_dir() and depth > 1:
                try:
                    entry["children"] = _scan_dir_recursive(child, depth - 1, include_hidden)
                except PermissionError:
                    entry["children"] = []
            entries.append(entry)
    except PermissionError:
        pass
    except Exception:
        pass
    return entries


def list_laptop_structure(
    include_hidden: bool = False,
    output_file: str = "",
    depth: int = 2,
) -> Dict[str, Any]:
    """
    Deterministic full-laptop scan: discovers all available drives and lists
    the contents of each drive root PLUS the major user directories
    (Home, Downloads, Desktop, Documents, Pictures, Music, Videos).

    For user-created top-level drive folders (e.g. C:\\Hrishikesh) the scan
    recurses up to *depth* levels so nested files are visible.

    This is the authoritative tool for "what files/folders are on my laptop"
    requests — it does NOT rely on an LLM to decide which folders to scan.

    Args:
        include_hidden: Include hidden files/folders (default False).
        output_file:    If provided, write the plain-text report to this path
                        and return it as file_path in the result.
        depth:          How many levels deep to recurse for user-created folders
                        at the drive root and user directories (default 2).
    """
    try:
        # ── 1. Discover drives ──────────────────────────────────────────────
        drives: List[Path] = []
        if platform.system() == "Windows":
            for letter in string.ascii_uppercase:
                drive_path = Path(f"{letter}:\\")
                try:
                    if drive_path.exists():
                        drives.append(drive_path)
                except Exception:
                    pass
        else:
            drives = [Path("/")]

        # ── 2. Scan root of each drive ──────────────────────────────────────
        # System folders get depth=1 (shallow); user-created folders recurse.
        drive_trees: Dict[str, Any] = {}
        for drive in drives:
            try:
                top_level = []
                for child in sorted(drive.iterdir(), key=lambda x: x.name.lower()):
                    if not include_hidden and (child.name.startswith(".") or child.name.startswith("$")):
                        continue
                    is_system = child.name in _SYSTEM_FOLDERS or not child.is_dir()
                    entry: Dict[str, Any] = {
                        "name": child.name,
                        "type": "folder" if child.is_dir() else "file",
                    }
                    if child.is_dir() and not is_system and depth > 1:
                        try:
                            entry["children"] = _scan_dir_recursive(child, depth - 1, include_hidden)
                        except PermissionError:
                            entry["children"] = []
                    top_level.append(entry)
                drive_trees[str(drive)] = top_level
            except PermissionError:
                drive_trees[str(drive)] = {"error": "Permission denied"}
            except Exception as exc:
                drive_trees[str(drive)] = {"error": str(exc)}

        # ── 3. Scan well-known user directories ────────────────────────────
        home = Path.home()
        user_dir_map: Dict[str, Path] = {
            "Home": home,
            "Downloads": home / "Downloads",
            "Desktop": home / "Desktop",
            "Documents": home / "Documents",
            "Pictures": home / "Pictures",
            "Music": home / "Music",
            "Videos": home / "Videos",
        }
        user_contents: Dict[str, Any] = {}
        for name, path in user_dir_map.items():
            if not path.exists():
                user_contents[name] = {"error": "Not found", "path": str(path)}
                continue
            try:
                entries = _scan_dir_recursive(path, depth, include_hidden)
                user_contents[name] = {
                    "path": str(path),
                    "total": len(entries),
                    "entries": entries,
                }
            except PermissionError:
                user_contents[name] = {"error": "Permission denied", "path": str(path)}
            except Exception as exc:
                user_contents[name] = {"error": str(exc), "path": str(path)}

        # ── 4. Build plain-text report ──────────────────────────────────────
        def _render_entries(entries: List[Dict[str, Any]], indent: int = 2) -> List[str]:
            """Recursively render a list of file/folder entries as text lines."""
            result_lines: List[str] = []
            pad = " " * indent
            for item in entries:
                marker = "[DIR] " if item["type"] == "folder" else "[FILE]"
                result_lines.append(f"{pad}{marker} {item['name']}")
                children = item.get("children")
                if children:
                    result_lines.extend(_render_entries(children, indent + 2))
            return result_lines

        lines = [
            f"LAPTOP STRUCTURE SCAN — {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
            f"Scan depth: {depth} level(s) inside user-created folders",
            "=" * 60,
        ]
        lines.append("\n=== DRIVES ===")
        for drive_path, contents in drive_trees.items():
            lines.append(f"\n{drive_path}")
            if isinstance(contents, dict) and "error" in contents:
                lines.append(f"  [Error: {contents['error']}]")
            else:
                lines.extend(_render_entries(contents, indent=2))

        lines.append("\n=== USER DIRECTORIES ===")
        for name, data in user_contents.items():
            if isinstance(data, dict) and "error" in data:
                lines.append(f"\n{name}: {data.get('path', '')} [Error: {data['error']}]")
            else:
                lines.append(f"\n{name}: {data['path']} ({data['total']} entries)")
                lines.extend(_render_entries(data.get("entries", []), indent=2))

        report_text = "\n".join(lines)

        result: Dict[str, Any] = {
            "status": "success",
            "drives": list(drive_trees.keys()),
            "drive_trees": drive_trees,
            "user_directories": user_contents,
            "report": report_text,
            "total_drives": len(drives),
        }

        # ── 5. Write to file — auto-generate path when caller omits output_file ──
        # ALWAYS write a local file so the artifact collector can surface it
        # for email attachment / Telegram delivery without the LLM having to
        # remember to pass output_file explicitly.
        if not output_file:
            reports_dir = Path.home() / "_octamind_reports"
            reports_dir.mkdir(parents=True, exist_ok=True)
            output_file = str(reports_dir / f"laptop_structure_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt")

        try:
            out_path = resolve_path(output_file)
            out_path.parent.mkdir(parents=True, exist_ok=True)
            out_path.write_text(report_text, encoding="utf-8")
            result["file_path"] = str(out_path)
            result["message"] = f"Laptop structure report written to {out_path} ({len(report_text)} chars)"
        except Exception as exc:
            result["file_write_error"] = str(exc)
            result["message"] = f"Report generated ({len(report_text)} chars) but could not write file: {exc}"

        return result

    except Exception as exc:
        logger.error("list_laptop_structure failed: %s", exc)
        return {"status": "error", "message": str(exc)}


def deliver_file(path: str) -> Dict[str, Any]:
    """
    Explicitly mark a local file for download delivery to the user.

    Call this once you have found or created the file the user asked for.
    Setting ``file_path`` in the return value causes the auto-responder
    to send it as a Telegram document or Dashboard download button.

    Args:
        path: Absolute (or ~-relative) path to the file.
    """
    try:
        p = Path(path).expanduser().resolve()
        if not p.exists():
            return {"status": "error", "message": f"File not found: {p}"}
        if not p.is_file():
            return {"status": "error", "message": f"Path is not a file: {p}"}
        stat = p.stat()
        size_str = (
            f"{stat.st_size / 1024:.1f} KB" if stat.st_size < 1_048_576
            else f"{stat.st_size / 1_048_576:.1f} MB"
        )
        return {
            "status": "success",
            "file_path": str(p),
            "filename": p.name,
            "size": size_str,
            "message": f"File ready: {p.name} ({size_str})",
        }
    except Exception as exc:
        return {"status": "error", "message": f"deliver_file error: {exc}"}


def write_pdf_report(path: str, title: str, content: str) -> Dict[str, Any]:
    """
    Write a formatted PDF report to *path*.

    Requires ``fpdf2`` (``pip install fpdf2``).  Falls back to writing a
    plain ``.txt`` file if the library is not available.

    Args:
        path:    Destination file path (should end in .pdf).
        title:   Report title printed at the top of the first page.
        content: Plain-text body (newlines respected).
    """
    try:
        out = Path(path).expanduser()
        out.parent.mkdir(parents=True, exist_ok=True)

        try:
            from fpdf import FPDF  # type: ignore
            pdf = FPDF()
            pdf.set_auto_page_break(auto=True, margin=15)
            pdf.add_page()
            pdf.set_font("Helvetica", "B", 16)
            pdf.cell(0, 10, title, ln=True, align="C")
            pdf.set_font("Helvetica", size=10)
            pdf.ln(4)
            for line in content.split("\n"):
                # multi_cell wraps long lines
                pdf.multi_cell(0, 6, line if line.strip() else " ")
            pdf.output(str(out))
        except ImportError:
            # Graceful fallback: write as annotated text file
            fallback = out.with_suffix(".txt")
            fallback.write_text(f"{title}\n{'=' * len(title)}\n\n{content}", encoding="utf-8")
            return {
                "status": "success",
                "file_path": str(fallback),
                "message": f"fpdf2 not installed — saved as plain text: {fallback.name}",
                "note": "Run 'pip install fpdf2' to enable real PDF generation.",
            }

        return {
            "status": "success",
            "file_path": str(out),
            "message": f"PDF report written to {out.name} ({out.stat().st_size} bytes)",
        }
    except Exception as exc:
        return {"status": "error", "message": f"write_pdf_report error: {exc}"}


def write_excel_report(
    path: str,
    sheet_data: Dict[str, Any],
    title: str = "",
) -> Dict[str, Any]:
    """
    Write a structured Excel workbook to *path*.

    Requires ``openpyxl`` (``pip install openpyxl``).

    Args:
        path:       Destination ``.xlsx`` file path.
        sheet_data: Dict mapping sheet name → list of row dicts OR list of
                    lists.  Example::

                        {"Summary": [{"Name": "Alice", "Score": 95}, ...]}

        title:      Optional title row inserted at the top of every sheet.
    """
    try:
        out = Path(path).expanduser()
        out.parent.mkdir(parents=True, exist_ok=True)

        try:
            import openpyxl  # type: ignore
            from openpyxl.styles import Font, PatternFill  # type: ignore
        except ImportError:
            return {
                "status": "error",
                "message": "openpyxl is not installed. Run: pip install openpyxl",
            }

        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # remove default empty sheet

        for sheet_name, rows in sheet_data.items():
            ws = wb.create_sheet(title=sheet_name[:31])  # Excel sheet name max 31 chars
            row_offset = 1

            if title:
                ws.cell(row=1, column=1, value=title).font = Font(bold=True, size=13)
                row_offset = 3

            if not rows:
                continue

            # Normalise: list-of-dicts → header row + data rows
            if isinstance(rows[0], dict):
                headers = list(rows[0].keys())
                header_row = ws[row_offset]
                for col_idx, h in enumerate(headers, 1):
                    cell = ws.cell(row=row_offset, column=col_idx, value=h)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill("solid", fgColor="D9E1F2")
                for r_idx, row_dict in enumerate(rows, row_offset + 1):
                    for col_idx, key in enumerate(headers, 1):
                        ws.cell(row=r_idx, column=col_idx, value=row_dict.get(key, ""))
            else:
                # list-of-lists
                for r_idx, row_list in enumerate(rows, row_offset):
                    for col_idx, val in enumerate(row_list, 1):
                        ws.cell(row=r_idx, column=col_idx, value=val)

        if out.suffix.lower() != ".xlsx":
            out = out.with_suffix(".xlsx")
        wb.save(str(out))
        return {
            "status": "success",
            "file_path": str(out),
            "sheets": list(sheet_data.keys()),
            "message": f"Excel workbook written to {out.name} ({out.stat().st_size} bytes)",
        }
    except Exception as exc:
        return {"status": "error", "message": f"write_excel_report error: {exc}"}


def organize_folder(
    directory: str,
    by: str = "extension",
    dry_run: bool = True,
    include_hidden: bool = False,
) -> Dict[str, Any]:
    """
    Organise files in *directory* by grouping them into sub-folders.

    The function ALWAYS does a dry-run first (``dry_run=True``) so the user
    can review the plan before files are actually moved.

    Args:
        directory:      Folder to organise.
        by:             Grouping strategy:
                        ``"extension"``  — group by file type (PDF/, Images/, etc.)
                        ``"date"``       — group by year-month of last modification
                        ``"name"``       — group A-Z by first letter of filename
                        ``"size"``       — group into Small / Medium / Large buckets
        dry_run:        When True (default) return the plan without moving anything.
                        Pass False to actually move the files.
        include_hidden: Include hidden files in the operation.
    """
    _EXT_GROUP: Dict[str, str] = {
        # Documents
        "pdf": "PDF", "doc": "Word", "docx": "Word", "txt": "Text",
        "odt": "Text", "rtf": "Text", "md": "Text",
        # Spreadsheets
        "xls": "Excel", "xlsx": "Excel", "csv": "CSV", "ods": "Spreadsheets",
        # Presentations
        "ppt": "Presentations", "pptx": "Presentations",
        # Images
        "jpg": "Images", "jpeg": "Images", "png": "Images",
        "gif": "Images", "bmp": "Images", "svg": "Images", "webp": "Images",
        # Audio
        "mp3": "Audio", "wav": "Audio", "flac": "Audio", "aac": "Audio",
        # Video
        "mp4": "Video", "avi": "Video", "mkv": "Video", "mov": "Video",
        # Archives
        "zip": "Archives", "rar": "Archives", "7z": "Archives", "tar": "Archives",
        # Code
        "py": "Code", "js": "Code", "ts": "Code", "html": "Code",
        "css": "Code", "json": "Code", "xml": "Code",
        # Executables / installers
        "exe": "Executables", "msi": "Executables", "sh": "Executables",
    }

    try:
        root = resolve_path(directory)
        if not root.exists() or not root.is_dir():
            return {"status": "error", "message": f"Directory not found: {root}"}

        # Collect files (non-recursive — only immediate children)
        files = [
            f for f in root.iterdir()
            if f.is_file()
            and (include_hidden or not f.name.startswith("."))
        ]
        if not files:
            return {"status": "success", "message": "No files to organise.", "plan": []}

        # Build move plan
        plan: list = []
        for f in sorted(files, key=lambda x: x.name.lower()):
            if by == "extension":
                ext = f.suffix.lstrip(".").lower()
                group = _EXT_GROUP.get(ext, f"Other_{ext.upper()}" if ext else "No_Extension")
            elif by == "date":
                mtime = datetime.fromtimestamp(f.stat().st_mtime)
                group = mtime.strftime("%Y-%m")
            elif by == "name":
                first = f.name[0].upper() if f.name else "#"
                group = first if first.isalpha() else "#"
            elif by == "size":
                size = f.stat().st_size
                if size < 1_048_576:        group = "Small_under1MB"
                elif size < 104_857_600:    group = "Medium_1MB_100MB"
                else:                       group = "Large_over100MB"
            else:
                return {"status": "error", "message": f"Unknown grouping: '{by}'. Use extension|date|name|size."}

            dest_dir = root / group
            dest_file = dest_dir / f.name
            plan.append({"file": f.name, "destination_folder": group, "full_dest": str(dest_file)})

        if dry_run:
            # Summarise plan by group
            groups: Dict[str, int] = {}
            for item in plan:
                groups[item["destination_folder"]] = groups.get(item["destination_folder"], 0) + 1
            summary_lines = [f"  {g}: {n} file(s)" for g, n in sorted(groups.items())]
            return {
                "status": "success",
                "dry_run": True,
                "directory": str(root),
                "total_files": len(files),
                "plan": plan,
                "groups": groups,
                "message": (
                    f"DRY RUN — {len(files)} files would be organised into {len(groups)} folders:\n"
                    + "\n".join(summary_lines)
                    + "\n\nCall again with dry_run=False to apply."
                ),
            }

        # Actually move
        moved, errors = [], []
        for item in plan:
            src = root / item["file"]
            dest_dir = root / item["destination_folder"]
            dest_dir.mkdir(exist_ok=True)
            dest = dest_dir / item["file"]
            try:
                src.rename(dest)
                moved.append(item["file"])
            except Exception as exc:
                errors.append({"file": item["file"], "error": str(exc)})

        return {
            "status": "success" if not errors else "partial",
            "moved": len(moved),
            "errors": errors,
            "message": f"Organised {len(moved)} file(s). {len(errors)} error(s).",
        }
    except Exception as exc:
        return {"status": "error", "message": f"organize_folder error: {exc}"}


# ─────────────────────────────────────────────────────────────────────────────
# Disk analysis
# ─────────────────────────────────────────────────────────────────────────────

def analyze_disk_usage(
    path: str,
    depth: int = 2,
    top_n: int = 20,
) -> Dict[str, Any]:
    """Recursively compute folder sizes and rank the biggest space consumers.

    Args:
        path:  Root directory to analyse.
        depth: How many levels deep to show (default 2).
        top_n: Number of largest items to surface in the summary.

    Returns:
        Dict with total_size, top folders, and a nested tree.
    """
    def _dir_size(p: Path) -> int:
        total = 0
        try:
            for entry in p.rglob("*"):
                try:
                    if entry.is_file(follow_symlinks=False):
                        total += entry.stat().st_size
                except (PermissionError, OSError):
                    pass
        except (PermissionError, OSError):
            pass
        return total

    def _walk(p: Path, cur_depth: int) -> Dict[str, Any]:
        node: Dict[str, Any] = {"name": p.name, "path": str(p), "type": "dir"}
        try:
            children = sorted(
                [c for c in p.iterdir() if not c.name.startswith(".")],
                key=lambda x: x.name.lower(),
            )
        except PermissionError:
            node["size_bytes"] = 0
            node["size_human"] = "—"
            node["error"] = "permission denied"
            return node

        child_nodes = []
        total = 0
        for c in children:
            if c.is_symlink():
                continue
            if c.is_dir():
                if cur_depth > 0:
                    child_node = _walk(c, cur_depth - 1)
                    child_nodes.append(child_node)
                    total += child_node.get("size_bytes", 0)
                else:
                    sz = _dir_size(c)
                    total += sz
                    child_nodes.append({
                        "name": c.name, "path": str(c), "type": "dir",
                        "size_bytes": sz, "size_human": _human_size(sz),
                    })
            else:
                try:
                    sz = c.stat().st_size
                    total += sz
                    child_nodes.append({
                        "name": c.name, "path": str(c), "type": "file",
                        "size_bytes": sz, "size_human": _human_size(sz),
                    })
                except (PermissionError, OSError):
                    pass

        node["size_bytes"] = total
        node["size_human"] = _human_size(total)
        node["children"] = sorted(child_nodes, key=lambda x: x.get("size_bytes", 0), reverse=True)
        return node

    try:
        root = resolve_path(path)
        if not root.exists():
            return {"status": "error", "message": f"Path not found: {root}"}

        tree = _walk(root, depth)
        total_bytes = tree["size_bytes"]

        # Flatten to find top N by size
        flat: list = []

        def _flatten(node: Dict[str, Any]) -> None:
            if node["type"] == "dir":
                flat.append(node)
            for child in node.get("children", []):
                _flatten(child)

        _flatten(tree)
        flat.sort(key=lambda x: x.get("size_bytes", 0), reverse=True)
        top = flat[:top_n]

        return {
            "status": "success",
            "root": str(root),
            "total_size_bytes": total_bytes,
            "total_size_human": _human_size(total_bytes),
            "top_consumers": [
                {"path": x["path"], "size": x["size_human"]} for x in top
            ],
            "tree": tree,
            "message": (
                f"Total size of '{root}': {_human_size(total_bytes)}. "
                f"Top space consumer: {top[0]['path']} ({top[0]['size_human']})" if top else ""
            ),
        }
    except Exception as exc:
        return {"status": "error", "message": f"analyze_disk_usage error: {exc}"}


def get_drive_info() -> Dict[str, Any]:
    """Return information about all available local drives (Windows) or mount
    points (Linux/macOS) including total, used and free space.

    Returns:
        Dict with per-drive stats and a combined total.
    """
    import shutil
    drives_info = []
    if platform.system() == "Windows":
        for letter in string.ascii_uppercase:
            dp = Path(f"{letter}:\\")
            try:
                if not dp.exists():
                    continue
                usage = shutil.disk_usage(dp)
                drives_info.append({
                    "drive": str(dp),
                    "total_bytes": usage.total,
                    "used_bytes":  usage.used,
                    "free_bytes":  usage.free,
                    "total_human": f"{usage.total / 1_073_741_824:.1f} GB",
                    "used_human":  f"{usage.used  / 1_073_741_824:.1f} GB",
                    "free_human":  f"{usage.free  / 1_073_741_824:.1f} GB",
                    "used_pct":    round(usage.used * 100 / usage.total, 1) if usage.total else 0,
                })
            except Exception:
                pass
    else:
        import subprocess
        try:
            out = subprocess.check_output(["df", "-k"], text=True)
            for line in out.splitlines()[1:]:
                parts = line.split()
                if len(parts) >= 6:
                    total = int(parts[1]) * 1024
                    used  = int(parts[2]) * 1024
                    free  = int(parts[3]) * 1024
                    drives_info.append({
                        "drive": parts[5],
                        "total_bytes": total, "used_bytes": used, "free_bytes": free,
                        "total_human": f"{total / 1_073_741_824:.1f} GB",
                        "used_human":  f"{used  / 1_073_741_824:.1f} GB",
                        "free_human":  f"{free  / 1_073_741_824:.1f} GB",
                        "used_pct": round(used * 100 / total, 1) if total else 0,
                    })
        except Exception as exc:
            return {"status": "error", "message": str(exc)}

    total_free  = sum(d["free_bytes"]  for d in drives_info)
    total_used  = sum(d["used_bytes"]  for d in drives_info)
    total_total = sum(d["total_bytes"] for d in drives_info)
    return {
        "status": "success",
        "drives": drives_info,
        "summary": {
            "total_human": f"{total_total / 1_073_741_824:.1f} GB",
            "used_human":  f"{total_used  / 1_073_741_824:.1f} GB",
            "free_human":  f"{total_free  / 1_073_741_824:.1f} GB",
        },
        "message": (
            f"{len(drives_info)} drive(s) found. "
            f"Total: {total_total / 1_073_741_824:.1f} GB, "
            f"Used: {total_used / 1_073_741_824:.1f} GB, "
            f"Free: {total_free / 1_073_741_824:.1f} GB."
        ),
    }


# ─────────────────────────────────────────────────────────────────────────────
# Duplicate detection (local filesystem)
# ─────────────────────────────────────────────────────────────────────────────

def find_duplicate_files(
    directory: str,
    recursive: bool = True,
    min_size_bytes: int = 1024,
) -> Dict[str, Any]:
    """Find duplicate files in *directory* by MD5 hash comparison.

    Args:
        directory:       Root directory to scan.
        recursive:       Recurse into sub-directories (default True).
        min_size_bytes:  Skip files smaller than this (default 1 KB).

    Returns:
        Dict with groups of duplicate files and total wasted space.
    """
    import hashlib

    def _md5(p: Path) -> str:
        h = hashlib.md5()
        with open(p, "rb") as f:
            for chunk in iter(lambda: f.read(65536), b""):
                h.update(chunk)
        return h.hexdigest()

    try:
        root = resolve_path(directory)
        if not root.exists():
            return {"status": "error", "message": f"Directory not found: {root}"}

        # Group by (size) first — only hash files that share size
        size_map: Dict[int, list] = {}
        glob = root.rglob("*") if recursive else root.iterdir()
        for f in glob:
            try:
                if not f.is_file(follow_symlinks=False):
                    continue
                sz = f.stat().st_size
                if sz < min_size_bytes:
                    continue
                size_map.setdefault(sz, []).append(f)
            except (PermissionError, OSError):
                pass

        # Hash only size-collision groups
        hash_map: Dict[str, list] = {}
        for sz, files in size_map.items():
            if len(files) < 2:
                continue
            for f in files:
                try:
                    h = _md5(f)
                    hash_map.setdefault(h, []).append(str(f))
                except Exception:
                    pass

        groups = [paths for paths in hash_map.values() if len(paths) > 1]
        wasted = sum(
            Path(paths[0]).stat().st_size * (len(paths) - 1)
            for paths in groups
            if Path(paths[0]).exists()
        )

        return {
            "status": "success",
            "duplicate_groups": len(groups),
            "total_duplicates": sum(len(g) - 1 for g in groups),
            "wasted_bytes": wasted,
            "wasted_human": _human_size(wasted) if wasted else "0 B",
            "groups": [{"count": len(g), "size": _human_size(Path(g[0]).stat().st_size), "files": g} for g in groups],
            "message": (
                f"Found {len(groups)} group(s) of duplicates "
                f"({sum(len(g) - 1 for g in groups)} extra copies) wasting "
                f"{_human_size(wasted)}."
            ),
        }
    except Exception as exc:
        return {"status": "error", "message": f"find_duplicate_files error: {exc}"}


# ─────────────────────────────────────────────────────────────────────────────
# Content search
# ─────────────────────────────────────────────────────────────────────────────

def search_files_by_content(
    query: str,
    directory: str = "~",
    extensions: Optional[List[str]] = None,
    max_results: int = 50,
    case_sensitive: bool = False,
) -> Dict[str, Any]:
    """Search files whose content contains *query* (grep-like).

    Args:
        query:          Text to search for inside files.
        directory:      Root directory to search (default: Home folder).
        extensions:     Restrict to these file extensions e.g. ['txt', 'py', 'md'].
                        If None, searches all text-like files.
        max_results:    Stop after this many matches.
        case_sensitive: Default False.

    Returns:
        Dict with matching files and line snippets.
    """
    _TEXT_EXTS = {
        "txt", "md", "py", "js", "ts", "json", "yaml", "yml", "csv",
        "html", "htm", "css", "xml", "sh", "bat", "ini", "cfg", "log",
        "rst", "toml", "env", "conf",
    }
    try:
        root = resolve_path(directory)
        allowed_exts = set(e.lstrip(".").lower() for e in extensions) if extensions else _TEXT_EXTS
        flag = 0 if case_sensitive else re.IGNORECASE
        pattern = re.compile(re.escape(query), flag)

        matches = []
        for f in root.rglob("*"):
            if len(matches) >= max_results:
                break
            try:
                if not f.is_file():
                    continue
                if f.suffix.lstrip(".").lower() not in allowed_exts:
                    continue
                if f.stat().st_size > 10_485_760:  # skip > 10 MB files
                    continue
                text = f.read_text(encoding="utf-8", errors="ignore")
                hit_lines = []
                for i, line in enumerate(text.splitlines(), 1):
                    if pattern.search(line):
                        hit_lines.append({"line": i, "text": line.strip()[:200]})
                if hit_lines:
                    matches.append({
                        "file": str(f),
                        "hit_count": len(hit_lines),
                        "lines": hit_lines[:5],  # show first 5 hits per file
                    })
            except Exception:
                pass

        return {
            "status": "success",
            "query": query,
            "directory": str(root),
            "match_count": len(matches),
            "matches": matches,
            "message": f"Found '{query}' in {len(matches)} file(s) under '{root}'.",
        }
    except Exception as exc:
        return {"status": "error", "message": f"search_files_by_content error: {exc}"}


# ─────────────────────────────────────────────────────────────────────────────
# Batch rename
# ─────────────────────────────────────────────────────────────────────────────

def batch_rename(
    directory: str,
    find: str,
    replace: str,
    dry_run: bool = True,
    use_regex: bool = False,
    extensions: Optional[List[str]] = None,
) -> Dict[str, Any]:
    """Rename all files in *directory* by replacing *find* with *replace*.

    Args:
        directory:  Directory containing files to rename.
        find:       Substring (or regex if use_regex=True) to find in filename.
        replace:    Replacement string.
        dry_run:    Show plan without actually renaming (default True).
        use_regex:  Treat *find* as a Python regex pattern.
        extensions: Only rename files with these extensions.

    Returns:
        Dict with rename plan and results.
    """
    try:
        root = resolve_path(directory)
        if not root.exists() or not root.is_dir():
            return {"status": "error", "message": f"Directory not found: {root}"}

        allowed_exts = set(e.lstrip(".").lower() for e in extensions) if extensions else None
        flag = re.compile(find) if use_regex else None

        plan = []
        for f in sorted(root.iterdir(), key=lambda x: x.name.lower()):
            if not f.is_file():
                continue
            if allowed_exts and f.suffix.lstrip(".").lower() not in allowed_exts:
                continue
            if use_regex and flag:
                new_name = flag.sub(replace, f.name)
            else:
                new_name = f.name.replace(find, replace)
            if new_name != f.name:
                plan.append({"old": f.name, "new": new_name, "path": str(f)})

        if dry_run:
            return {
                "status": "success",
                "dry_run": True,
                "rename_count": len(plan),
                "plan": plan,
                "message": (
                    f"DRY RUN — {len(plan)} file(s) would be renamed. "
                    "Call again with dry_run=False to apply."
                ),
            }

        renamed, errors = [], []
        for item in plan:
            src  = Path(item["path"])
            dest = src.parent / item["new"]
            if dest.exists():
                errors.append({"old": item["old"], "error": f"'{item['new']}' already exists"})
                continue
            try:
                src.rename(dest)
                renamed.append(item)
            except Exception as exc:
                errors.append({"old": item["old"], "error": str(exc)})

        return {
            "status": "success" if not errors else "partial",
            "renamed": len(renamed),
            "errors": errors,
            "message": f"Renamed {len(renamed)} file(s). {len(errors)} error(s).",
        }
    except Exception as exc:
        return {"status": "error", "message": f"batch_rename error: {exc}"}


# ─────────────────────────────────────────────────────────────────────────────
# Secure delete
# ─────────────────────────────────────────────────────────────────────────────

def secure_delete(path: str, passes: int = 3) -> Dict[str, Any]:
    """Overwrite a file with random bytes N times before deleting it.
    This makes recovery significantly harder for sensitive files.

    Args:
        path:   Path to the file to securely delete.
        passes: Number of overwrite passes (default 3).

    Returns:
        Dict with status.
    """
    import os
    try:
        target = resolve_path(path)
        if not target.is_file():
            return {"status": "error", "message": f"File not found or is a directory: {target}"}
        size = target.stat().st_size
        with open(target, "r+b") as fh:
            for _ in range(passes):
                fh.seek(0)
                fh.write(os.urandom(size))
                fh.flush()
                os.fsync(fh.fileno())
        target.unlink()
        return {
            "status": "success",
            "path": str(target),
            "passes": passes,
            "message": f"'{target.name}' securely deleted ({passes} overwrite pass(es)).",
        }
    except Exception as exc:
        return {"status": "error", "message": f"secure_delete error: {exc}"}


# ─────────────────────────────────────────────────────────────────────────────
# Temp file cleanup
# ─────────────────────────────────────────────────────────────────────────────

def cleanup_temp_files(dry_run: bool = True) -> Dict[str, Any]:
    """Safely remove Windows / macOS / Linux temporary files and caches.

    Targets:
      - %TEMP% / /tmp  (OS temp folder)
      - Recycle Bin    ($Recycle.Bin on Windows)
      - Windows Update Cache (%WinDir%\\SoftwareDistribution\\Download)
      - Python __pycache__ in the current project

    Args:
        dry_run: Show what would be removed without actually deleting (default True).

    Returns:
        Dict with list of targets, sizes and result.
    """
    import shutil

    targets = []

    # OS temp folder
    import tempfile
    tmp = Path(tempfile.gettempdir())
    if tmp.exists():
        targets.append(("OS Temp", tmp, False))  # (label, path, skip_errors)

    if platform.system() == "Windows":
        # Windows Update download cache
        windir = os.environ.get("WINDIR", "C:\\Windows")
        wu_cache = Path(windir) / "SoftwareDistribution" / "Download"
        if wu_cache.exists():
            targets.append(("Windows Update Cache", wu_cache, True))

    # Python __pycache__ in CWD project tree
    cwd = Path.cwd()
    pycache_dirs = list(cwd.rglob("__pycache__"))
    for pc in pycache_dirs:
        targets.append(("__pycache__", pc, True))

    results = []
    total_freed = 0

    for label, p, skip_errors in targets:
        try:
            size = sum(f.stat().st_size for f in p.rglob("*") if f.is_file())
        except Exception:
            size = 0

        if dry_run:
            results.append({
                "label": label, "path": str(p),
                "size_human": _human_size(size), "action": "would delete",
            })
            total_freed += size
        else:
            try:
                if p.is_dir():
                    for child in p.iterdir():
                        try:
                            if child.is_dir():
                                shutil.rmtree(child, ignore_errors=skip_errors)
                            else:
                                child.unlink(missing_ok=True)
                        except Exception:
                            if not skip_errors:
                                raise
                results.append({
                    "label": label, "path": str(p),
                    "size_human": _human_size(size), "action": "cleaned",
                })
                total_freed += size
            except Exception as exc:
                results.append({
                    "label": label, "path": str(p),
                    "size_human": _human_size(size), "action": f"error: {exc}",
                })

    verb = "would free" if dry_run else "freed"
    return {
        "status": "success",
        "dry_run": dry_run,
        "targets": results,
        "total_freed_human": _human_size(total_freed),
        "message": (
            f"{'DRY RUN — ' if dry_run else ''}"
            f"{len(results)} temp location(s) scanned. {verb} ~{_human_size(total_freed)}. "
            + ("Call with dry_run=False to apply." if dry_run else "Done!")
        ),
    }


# ─────────────────────────────────────────────────────────────────────────────
# Folder monitor
# ─────────────────────────────────────────────────────────────────────────────

def monitor_folder(
    path: str,
    timeout_seconds: int = 30,
    poll_interval: float = 1.0,
) -> Dict[str, Any]:
    """Watch a folder for new, modified or deleted files for up to *timeout_seconds*.

    This is a lightweight polling monitor (no external libraries required).
    Returns the list of changes detected.

    Args:
        path:             Folder to watch.
        timeout_seconds:  How long to watch (default 30 s).
        poll_interval:    Check interval in seconds (default 1.0 s).

    Returns:
        Dict with list of detected changes.
    """
    import time

    try:
        root = resolve_path(path)
        if not root.exists() or not root.is_dir():
            return {"status": "error", "message": f"Directory not found: {root}"}

        def _snapshot(p: Path) -> Dict[str, float]:
            snap: Dict[str, float] = {}
            for f in p.rglob("*"):
                try:
                    snap[str(f)] = f.stat().st_mtime
                except (PermissionError, OSError):
                    pass
            return snap

        before = _snapshot(root)
        start  = time.monotonic()
        changes: list = []

        while time.monotonic() - start < timeout_seconds:
            time.sleep(poll_interval)
            after = _snapshot(root)

            # New files
            for fp in set(after) - set(before):
                changes.append({"event": "created", "path": fp})
            # Deleted files
            for fp in set(before) - set(after):
                changes.append({"event": "deleted", "path": fp})
            # Modified files
            for fp in set(before) & set(after):
                if after[fp] != before[fp]:
                    changes.append({"event": "modified", "path": fp})

            if changes:
                break
            before = after

        return {
            "status": "success",
            "folder": str(root),
            "watched_seconds": round(time.monotonic() - start, 1),
            "changes": changes,
            "change_count": len(changes),
            "message": (
                f"{len(changes)} change(s) detected in '{root}' "
                f"within {round(time.monotonic() - start, 1)}s."
                if changes else
                f"No changes detected in '{root}' after {timeout_seconds}s."
            ),
        }
    except Exception as exc:
        return {"status": "error", "message": f"monitor_folder error: {exc}"}


# ─────────────────────────────────────────────────────────────────────────────
# App cache cleanup
# ─────────────────────────────────────────────────────────────────────────────

def cleanup_app_caches(dry_run: bool = True) -> Dict[str, Any]:
    """Remove browser caches, installer temp files, and log archives safely.

    Targets on Windows: Chrome, Edge, Firefox caches, %LOCALAPPDATA%\\Temp.
    Targets on Linux/macOS: ~/.cache/google-chrome, ~/.mozilla/firefox, ~/.cache.

    Always use dry_run=True first to see what will be cleaned, then dry_run=False.
    """
    import sys as _sys, os as _os
    home = Path.home()
    targets: list = []  # list of (path, label)
    if _sys.platform == "win32":
        local_app = Path(_os.environ.get("LOCALAPPDATA", str(home / "AppData" / "Local")))
        chrome_cache = local_app / "Google" / "Chrome" / "User Data" / "Default" / "Cache"
        if chrome_cache.exists():
            targets.append((chrome_cache, "Chrome cache"))
        edge_cache = local_app / "Microsoft" / "Edge" / "User Data" / "Default" / "Cache"
        if edge_cache.exists():
            targets.append((edge_cache, "Edge cache"))
        ff_profiles_root = local_app / "Mozilla" / "Firefox" / "Profiles"
        if ff_profiles_root.exists():
            for profile_dir in ff_profiles_root.glob("*.default*"):
                ff_cache = profile_dir / "cache2"
                if ff_cache.exists():
                    targets.append((ff_cache, f"Firefox cache ({profile_dir.name})"))
        win_temp = local_app / "Temp"
        if win_temp.exists():
            targets.append((win_temp, "Windows AppData Temp"))
    else:
        chrome_cache = home / ".cache" / "google-chrome"
        if chrome_cache.exists():
            targets.append((chrome_cache, "Chrome cache"))
        ff_cache = home / ".mozilla" / "firefox"
        if ff_cache.exists():
            targets.append((ff_cache, "Firefox profile cache"))
        user_cache = home / ".cache"
        if user_cache.exists():
            for subdir in user_cache.iterdir():
                if subdir.is_dir() and subdir.name not in ("thumbnails",):
                    targets.append((subdir, f"~/.cache/{subdir.name}"))

    items = []
    total_size = 0
    for target_path, label in targets:
        dir_size = sum(
            f.stat().st_size for f in Path(target_path).rglob("*") if f.is_file()
        )
        total_size += dir_size
        items.append({
            "label": label,
            "path": str(target_path),
            "size_human": _human_size(dir_size),
            "action": "would delete" if dry_run else "deleted",
        })

    if not dry_run:
        import shutil as _shutil
        for target_path, _ in targets:
            try:
                _shutil.rmtree(target_path, ignore_errors=True)
            except Exception:
                pass

    return {
        "status": "dry_run" if dry_run else "success",
        "targets_found": len(items),
        "total_size_human": _human_size(total_size),
        "items": items,
        "message": (
            f"DRY RUN: Would free ~{_human_size(total_size)} from "
            f"{len(items)} cache location(s). Call with dry_run=False to delete."
            if dry_run else
            f"Cleaned {len(items)} cache location(s), freed ~{_human_size(total_size)}."
        ),
    }


# ─────────────────────────────────────────────────────────────────────────────
# Archive old files
# ─────────────────────────────────────────────────────────────────────────────

def archive_old_files(
    folder: str,
    months_old: int = 6,
    output_zip: str = "",
    dry_run: bool = True,
) -> Dict[str, Any]:
    """Find files not modified in N months and compress them into a zip archive.

    Always call with dry_run=True first (the default) to preview the plan.
    Then call again with dry_run=False to create the archive.

    Args:
        folder:     Directory to scan.
        months_old: Age threshold in months (default 6).
        output_zip: Output zip file path. Auto-generated if empty.
        dry_run:    Preview without creating archive (default True).
    """
    import datetime as _dt, zipfile as _zf
    try:
        folder_path = Path(folder).expanduser()
        if not folder_path.is_dir():
            return {"status": "error", "message": f"'{folder}' is not a directory."}
        cutoff = _dt.datetime.now() - _dt.timedelta(days=months_old * 30)
        old_files = [
            f for f in folder_path.rglob("*")
            if f.is_file() and _dt.datetime.fromtimestamp(f.stat().st_mtime) < cutoff
        ]
        if not old_files:
            return {
                "status": "success",
                "archived": 0,
                "message": f"No files older than {months_old} months found in '{folder}'.",
            }
        total_size = sum(f.stat().st_size for f in old_files)
        if dry_run:
            return {
                "status": "dry_run",
                "would_archive": len(old_files),
                "total_size_human": _human_size(total_size),
                "files": [str(f) for f in old_files[:50]],
                "message": (
                    f"DRY RUN: Would archive {len(old_files)} file(s) "
                    f"({_human_size(total_size)}). Call with dry_run=False to create the zip."
                ),
            }
        if not output_zip:
            ts = _dt.datetime.now().strftime("%Y%m%d_%H%M%S")
            output_zip = str(folder_path / f"archive_older_{months_old}mo_{ts}.zip")
        zip_path = Path(output_zip).expanduser()
        with _zf.ZipFile(zip_path, "w", compression=_zf.ZIP_DEFLATED) as zf:
            for f in old_files:
                zf.write(f, f.relative_to(folder_path))
        zip_size = zip_path.stat().st_size
        return {
            "status": "success",
            "archived": len(old_files),
            "original_size_human": _human_size(total_size),
            "zip_path": str(zip_path),
            "zip_size_human": _human_size(zip_size),
            "compression_ratio": f"{(1 - zip_size / max(total_size, 1)) * 100:.1f}%",
            "message": (
                f"Archived {len(old_files)} file(s) ({_human_size(total_size)}) "
                f"→ '{zip_path}' ({_human_size(zip_size)})."
            ),
        }
    except Exception as exc:
        return {"status": "error", "message": f"archive_old_files error: {exc}"}


# ─────────────────────────────────────────────────────────────────────────────
# Windows shortcut resolution
# ─────────────────────────────────────────────────────────────────────────────

def resolve_shortcut(lnk_path: str) -> Dict[str, Any]:
    """Resolve a Windows .lnk shortcut to its target path.

    Uses pywin32 (win32com) if available; falls back to binary parsing of the
    MS-SHLLINK format. On non-Windows returns an informative error.
    """
    import sys as _sys, struct as _st
    try:
        lnk = Path(lnk_path).expanduser()
        if not lnk.exists():
            return {"status": "error", "message": f"File not found: '{lnk_path}'"}
        if _sys.platform == "win32":
            try:
                import win32com.client as _wc  # pywin32
                shell = _wc.Dispatch("WScript.Shell")
                sc = shell.CreateShortCut(str(lnk))
                return {
                    "status": "success",
                    "shortcut_path": str(lnk),
                    "target_path": sc.TargetPath,
                    "working_directory": sc.WorkingDirectory,
                    "arguments": sc.Arguments,
                    "description": sc.Description,
                    "icon_location": sc.IconLocation,
                }
            except ImportError:
                pass  # fall through to binary parser
            # Minimal MS-SHLLINK binary parser
            data = lnk.read_bytes()
            hdr_size = _st.unpack_from("<I", data, 0)[0]
            link_flags = _st.unpack_from("<I", data, 0x14)[0]
            offset = hdr_size
            target = ""
            has_idlist = (link_flags & 0x01) != 0
            if has_idlist:
                idl_size = _st.unpack_from("<H", data, offset)[0]
                offset += 2 + idl_size
            has_linkinfo = (link_flags & 0x02) != 0
            if has_linkinfo:
                li_size = _st.unpack_from("<I", data, offset)[0]
                li_flags = _st.unpack_from("<I", data, offset + 4)[0]
                lbp_offset = _st.unpack_from("<I", data, offset + 16)[0]
                if li_flags & 0x01:
                    bp = offset + lbp_offset
                    end = data.index(b"\x00", bp)
                    target = data[bp:end].decode("latin-1", errors="ignore")
            return {
                "status": "success",
                "shortcut_path": str(lnk),
                "target_path": target or "(parsing incomplete — install pywin32 for full support)",
                "note": "Install pywin32 (`pip install pywin32`) for complete shortcut resolution.",
            }
        else:
            if lnk.suffix.lower() == ".lnk":
                return {"status": "error",
                        "message": ".lnk files are Windows format; cannot resolve on this OS."}
            return {"status": "success", "shortcut_path": str(lnk),
                    "target_path": str(lnk.resolve())}
    except Exception as exc:
        return {"status": "error", "message": f"resolve_shortcut error: {exc}"}


# ─────────────────────────────────────────────────────────────────────────────
# File hashing
# ─────────────────────────────────────────────────────────────────────────────

def get_file_hash(file_path: str, algorithm: str = "md5") -> Dict[str, Any]:
    """Compute the cryptographic hash of a local file.

    Args:
        file_path:  Path to the file.
        algorithm:  Hash algorithm: 'md5' (default), 'sha256', or 'sha1'.

    Useful for verifying file integrity and manually detecting duplicates.
    """
    import hashlib as _hl
    try:
        alg = algorithm.lower().replace("-", "")
        if alg not in ("md5", "sha256", "sha1"):
            return {"status": "error",
                    "message": f"Unsupported algorithm '{algorithm}'. Use: md5, sha256, sha1."}
        p = Path(file_path).expanduser()
        if not p.exists():
            return {"status": "error", "message": f"File not found: '{file_path}'"}
        if not p.is_file():
            return {"status": "error", "message": f"'{file_path}' is not a file."}
        h = _hl.new(alg)
        size = p.stat().st_size
        CHUNK = 1024 * 1024  # 1 MB
        with open(p, "rb") as fh:
            while chunk := fh.read(CHUNK):
                h.update(chunk)
        return {
            "status": "success",
            "path": str(p),
            "algorithm": algorithm.upper(),
            "hash": h.hexdigest(),
            "size_human": _human_size(size),
            "size_bytes": size,
        }
    except Exception as exc:
        return {"status": "error", "message": f"get_file_hash error: {exc}"}


# ─────────────────────────────────────────────────────────────────────────────
# Running processes
# ─────────────────────────────────────────────────────────────────────────────

def list_running_apps() -> Dict[str, Any]:
    """List currently running applications and processes on this machine.

    Returns process name, PID, memory usage, and CPU percentage.
    Uses psutil if available (richest data), falls back to tasklist / ps -aux.
    """
    try:
        import psutil as _pu  # type: ignore
        processes = []
        for proc in _pu.process_iter(["pid", "name", "exe", "memory_info", "cpu_percent", "status"]):
            try:
                info = proc.info
                if info.get("status") == "zombie":
                    continue
                mem = info.get("memory_info")
                processes.append({
                    "pid": info["pid"],
                    "name": info["name"],
                    "exe": info.get("exe") or "",
                    "memory_human": _human_size(mem.rss) if mem else "?",
                    "memory_bytes": mem.rss if mem else 0,
                    "cpu_pct": info.get("cpu_percent", 0.0),
                    "status": info.get("status", ""),
                })
            except (_pu.NoSuchProcess, _pu.AccessDenied):
                continue
        processes.sort(key=lambda x: x["memory_bytes"], reverse=True)
        return {
            "status": "success",
            "count": len(processes),
            "processes": processes,
            "message": f"{len(processes)} running process(es) found.",
        }
    except ImportError:
        pass
    # Fallback via OS command
    import subprocess as _sp, sys as _sys
    try:
        if _sys.platform == "win32":
            result = _sp.run(
                ["tasklist", "/fo", "csv", "/nh"],
                capture_output=True, text=True, timeout=15,
            )
            processes = []
            for line in result.stdout.strip().splitlines():
                parts = [p.strip('"') for p in line.split('","')]
                if len(parts) >= 2:
                    processes.append({
                        "name": parts[0],
                        "pid": parts[1],
                        "memory_human": parts[4] if len(parts) > 4 else "?",
                    })
        else:
            result = _sp.run(["ps", "aux"], capture_output=True, text=True, timeout=15)
            processes = []
            for line in result.stdout.strip().splitlines()[1:]:
                cols = line.split(None, 10)
                if len(cols) >= 11:
                    processes.append({
                        "name": cols[10], "pid": cols[1],
                        "cpu_pct": cols[2], "memory_pct": cols[3],
                    })
        return {
            "status": "success",
            "count": len(processes),
            "processes": processes,
            "note": "Install psutil for richer data: pip install psutil",
            "message": f"{len(processes)} running process(es) found.",
        }
    except Exception as exc:
        return {"status": "error", "message": f"list_running_apps error: {exc}"}


# ──────────────────────────────────────────────────────────────────────────────
# Manifest helpers — reliable cross-turn path bookkeeping
# Stored in <workspace>/data/ so the path is consistent on Windows/Linux/macOS.
# __file__ = src/files/features/file_ops.py
#   parents[0] = src/files/features/
#   parents[1] = src/files/
#   parents[2] = src/
#   parents[3] = <workspace root>
# ──────────────────────────────────────────────────────────────────────────────

_DEFAULT_OCTAMIND_DIR = Path(__file__).resolve().parents[3] / "data"
_DEFAULT_MANIFEST     = _DEFAULT_OCTAMIND_DIR / "octa_manifest.txt"
# Operation history (replaces single-entry last_operation.json).
# JSON array, most-recent-first.  Entries older than 30 days are pruned
# automatically on every write so the file stays small.
_OP_HISTORY_FILE      = _DEFAULT_OCTAMIND_DIR / "operation_history.json"
_OP_HISTORY_TTL_DAYS  = 30


def _prune_op_history(records: list) -> list:
    """Drop entries whose timestamp is older than _OP_HISTORY_TTL_DAYS."""
    cutoff = datetime.now() - timedelta(days=_OP_HISTORY_TTL_DAYS)
    return [
        r for r in records
        if datetime.fromisoformat(r.get("timestamp", "1970-01-01")) >= cutoff
    ]


def _load_op_history() -> list:
    """Read operation_history.json; return [] on any error or if missing."""
    import json as _json
    try:
        if _OP_HISTORY_FILE.exists():
            return _json.loads(_OP_HISTORY_FILE.read_text(encoding="utf-8"))
    except Exception:
        pass
    return []


def _save_op_history(records: list) -> None:
    """Write records to operation_history.json atomically."""
    import json as _json
    try:
        _OP_HISTORY_FILE.parent.mkdir(parents=True, exist_ok=True)
        _OP_HISTORY_FILE.write_text(
            _json.dumps(records, indent=2, ensure_ascii=False),
            encoding="utf-8",
        )
    except Exception:
        pass


def _log_operation(op_type: str, destination: str, count: int) -> None:
    """Push a copy/collect operation onto the audit history stack.

    The in-memory representation is most-recent-first.  Entries older than
    ``_OP_HISTORY_TTL_DAYS`` are pruned on every push so the file never grows
    unbounded.  Never raises — logging must not crash the calling operation.
    """
    try:
        history = _load_op_history()
        history.insert(0, {
            "type":        op_type,
            "destination": str(destination),
            "count":       count,
            "timestamp":   datetime.now().isoformat(timespec="seconds"),
            "undone":      False,
        })
        history = _prune_op_history(history)
        _save_op_history(history)
    except Exception:
        pass  # never let logging crash the main operation


def undo_last_file_operation() -> Dict[str, Any]:
    """Undo the most-recent undoable copy/collect operation.

    Reads the top entry from ``data/operation_history.json``, deletes the
    destination folder that was created, and marks the entry as ``undone``.
    The entry is **kept** in the history for auditing purposes; it will only
    be removed when it ages past 30 days.

    Returns:
        dict with keys: status, message, deleted_folder, count.
    """
    try:
        history = _load_op_history()
        if not history:
            return {
                "status":  "error",
                "message": "No previous operation found to undo.  "
                           "Run a copy/collect operation first.",
            }

        # Find the most-recent non-undone entry
        target_idx = next(
            (i for i, r in enumerate(history) if not r.get("undone", False)),
            None,
        )
        if target_idx is None:
            return {
                "status":  "error",
                "message": "All recent operations have already been undone.",
            }

        op  = history[target_idx]
        dst = Path(op["destination"])
        if not dst.exists():
            # Mark as undone even if folder is already gone
            history[target_idx]["undone"]    = True
            history[target_idx]["undone_at"] = datetime.now().isoformat(timespec="seconds")
            _save_op_history(history)
            return {
                "status":  "error",
                "message": f"Destination folder '{dst}' no longer exists — nothing to delete.",
            }

        shutil.rmtree(str(dst), ignore_errors=False)

        # Mark undone (preserve for audit) and persist
        history[target_idx]["undone"]    = True
        history[target_idx]["undone_at"] = datetime.now().isoformat(timespec="seconds")
        _save_op_history(history)

        return {
            "status":         "success",
            "deleted_folder": str(dst),
            "count":          op.get("count", 0),
            "message": (
                f"Undo complete — deleted '{dst}' ({op.get('count', 0)} file(s)).  "
                f"Operation recorded in audit log."
            ),
        }
    except Exception as exc:
        logger.error("undo_last_file_operation failed: %s", exc)
        return {"status": "error", "message": str(exc)}


def list_file_operations(days: int = _OP_HISTORY_TTL_DAYS) -> Dict[str, Any]:
    """
    Return the file-operation audit history for the last *days* days.

    Useful for reviewing what was copied/collected and whether operations
    have been undone.  Each entry contains: type, destination, count,
    timestamp, undone, undone_at (if applicable).

    Args:
        days:  How many days of history to include.  Defaults to 30.

    Returns:
        dict with keys: status, operations (list, newest-first), count.
    """
    try:
        history  = _load_op_history()
        cutoff   = datetime.now() - timedelta(days=days)
        filtered = [
            r for r in history
            if datetime.fromisoformat(r.get("timestamp", "1970-01-01")) >= cutoff
        ]
        return {
            "status":     "success",
            "operations": filtered,
            "count":      len(filtered),
            "message":    f"{len(filtered)} operation(s) in last {days} day(s).",
        }
    except Exception as exc:
        logger.error("list_file_operations failed: %s", exc)
        return {"status": "error", "message": str(exc)}


def save_search_manifest(
    found_paths: List[str],
    manifest_path: str = "",
) -> Dict[str, Any]:
    """
    Persist a list of found file paths to a plain-text manifest file
    (one absolute path per line).  Call this immediately after any search
    so the next user turn can copy/move/zip EVERY result, not just the few
    that fit inside the assistant reply.

    Args:
        found_paths:   List of absolute file paths returned by a search step.
        manifest_path: Where to write the manifest.  Defaults to
                       <workspace>/data/octa_manifest.txt
    """
    try:
        target = Path(manifest_path).expanduser() if manifest_path else _DEFAULT_MANIFEST
        target.parent.mkdir(parents=True, exist_ok=True)

        # Deduplicate while preserving order
        seen: set = set()
        unique: List[str] = []
        for p in found_paths:
            ps = str(p).strip()
            if ps and ps not in seen:
                seen.add(ps)
                unique.append(ps)

        target.write_text("\n".join(unique), encoding="utf-8")
        return {
            "status": "success",
            "manifest_path": str(target),
            "count": len(unique),
            "message": f"Saved {len(unique)} path(s) to manifest '{target}'.",
        }
    except Exception as exc:
        logger.error("save_search_manifest failed: %s", exc)
        return {"status": "error", "message": str(exc)}


def collect_files_from_manifest(
    manifest_path: str = "",
    destination: str = "",
) -> Dict[str, Any]:
    """
    Read file paths from a manifest file (created by save_search_manifest)
    and copy EVERY listed file into *destination*.

    This is the most reliable way to copy files found in a previous turn —
    it is NOT limited by how many paths fit in the session state.

    Args:
        manifest_path: Path to the manifest text file.  Defaults to
                       <workspace>/data/octa_manifest.txt
        destination:   Folder to copy files into.  Defaults to
                       <workspace>/data/
    """
    try:
        mf = Path(manifest_path).expanduser() if manifest_path else _DEFAULT_MANIFEST
        if not mf.exists():
            return {
                "status": "error",
                "message": f"Manifest file not found: '{mf}'.  Run a search first.",
            }

        lines = [ln.strip() for ln in mf.read_text(encoding="utf-8").splitlines()]
        file_paths = [ln for ln in lines if ln]

        if not file_paths:
            return {
                "status": "error",
                "message": f"Manifest file '{mf}' is empty.  Run a search first.",
            }

        dst_path = (
            Path(destination).expanduser() if destination else _DEFAULT_OCTAMIND_DIR
        )
        dst_path.mkdir(parents=True, exist_ok=True)

        copied: List[str] = []
        skipped: List[str] = []
        for src_str in file_paths:
            src = Path(src_str)
            if not src.exists():
                skipped.append(str(src))
                continue
            target = dst_path / src.name
            if target.exists():
                stem, suffix = src.stem, src.suffix
                for i in range(1, 10000):
                    target = dst_path / f"{stem}_{i}{suffix}"
                    if not target.exists():
                        break
            try:
                if src.is_dir():
                    # Merge the directory tree into destination using copytree.
                    # dirs_exist_ok=True (Python ≥3.8) is safe for incremental copies.
                    target_dir = dst_path / src.name
                    shutil.copytree(str(src), str(target_dir), dirs_exist_ok=True)
                    copied.append(str(target_dir))
                else:
                    shutil.copy2(str(src), str(target))
                    copied.append(str(target))
            except Exception as exc2:
                logger.warning("collect_files_from_manifest: skip '%s': %s", src, exc2)
                skipped.append(str(src))

        _log_operation("copy", str(dst_path), len(copied))
        return {
            "status": "success",
            "destination": str(dst_path),
            "file_path": str(dst_path),   # convenience alias for zip_folder
            "manifest_path": str(mf),
            "copied_count": len(copied),
            "copied": copied,
            "skipped": skipped,
            "message": (
                f"Copied {len(copied)} file(s) into '{dst_path}'."
                + (f"  Skipped {len(skipped)} missing/unreadable path(s)." if skipped else "")
            ),
        }
    except Exception as exc:
        logger.error("collect_files_from_manifest failed: %s", exc)
        return {"status": "error", "message": str(exc)}
